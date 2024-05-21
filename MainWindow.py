from PyQt5 import QtWidgets, uic, QtCore, QtGui
from pyqtgraph import PlotWidget, plot
from PyQt5.QtWidgets import QVBoxLayout,QMessageBox
import pandas as pd
import sys
import os
import res
import openpyxl
import PyQt5.QtWebEngineWidgets
from datetime import datetime
import serial
import random
import threading
import time

class Stopwatch():
    def __init__(self):
        self.StartTime = 0
        self.EndTime = 0
        self.TimePassed = 0

    def start(self):
        self.StartTime = time.time()

    def secondsPassed(self):
        self.EndTime = time.time()
        self.TimePassed = self.EndTime - self.StartTime
        return self.TimePassed
    
    def reset(self):
        self.StartTime = 0
        self.EndTime = 0
        self.TimePassed = 0

class UserInfo():
	def __init__(self,username= None,account= None):
		self.username = username
		if account==1:
			self.type= "Admin"
		elif account==2:
			self.type= "Patient"
		elif account==3:
			self.type= "Nurse"

	def setType(account:int):
		if account==1:
			self.type= "Admin"
		elif account==2:
			self.type= "Patient"
		elif account==3:
			self.type= "Nurse"

 
	def run(self):
		self.isRunning = True
		if self._target is not None:
			self._return = self._target(*self._args, **self._kwargs)
		self.isRunning = False
		self.returnValue = self._return

session = UserInfo()
mode = 0

class LoginWindow(QtWidgets.QMainWindow):
	def __init__(self):
		super().__init__()
		self.ui = uic.loadUi("login.ui", self)
		self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
		self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
		self.ui.comboBox.setEditable(True)
		self.ui.comboBox.setCurrentIndex(-1)
		self.ui.comboBox.setCurrentText("Choose Account Type")
		self.ui.comboBox.addItem("Patient")
		self.ui.comboBox.addItem("Nurse")
		self.ui.comboBox.addItem("Admin")
		self.ui.stackedWidget.setCurrentIndex(0)
		self.isLogin = True

		self.ui.pushButton_12.clicked.connect(self.minimize)
		self.ui.pushButton_10.clicked.connect(self.minimize)
		self.ui.pushButton_9.clicked.connect(self.exit)
		self.ui.pushButton_14.clicked.connect(self.exit)
		self.ui.pushButton_2.clicked.connect(self.SwitchLogin)
		self.ui.pushButton_6.clicked.connect(self.SwitchLogin)
		self.ui.pushButton.clicked.connect(self.Login)
		self.ui.pushButton_5.clicked.connect(self.Register)
		self.ui.comboBox.currentIndexChanged.connect(self.UpdateCombo)
		self.ui.comboBox.currentTextChanged.connect(self.comboBoxFixer)

	def UpdateCombo(self):
		self.ui.comboBox.setEditable(False)

	def SwitchLogin(self):
		self.isLogin = not self.isLogin
		if self.isLogin:
			self.ui.stackedWidget.setCurrentIndex(0)
		else:
			self.ui.stackedWidget.setCurrentIndex(1)

	def Login(self):
		username = self.ui.lineEdit.text()
		password = self.ui.lineEdit_2.text()
		users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Admin','Patient','Nurse'])
		admins = users['Admin']
		patients = users['Patient']
		nurses = users['Nurse']
		found = False
		real_pw = ''
		c=0
		for user in [admins,patients,nurses]:
			if found:
				break
			c += 1
			for i,p in zip(user['username'],user['password']):
				if username == i:
					found = True
					real_pw = p
					break

		if found and (real_pw == password):
			self.ui.label_17.setText("Welcome!")
			global mode,session
			mode = c
			session = UserInfo(username=username,account=c)
			win = MainWindow()
			win.setWindowTitle("Hospital System")
			win.setWindowIcon(QtGui.QIcon("assets\\865969.png"))
			win.show()
			self.close()
		elif found and (real_pw != password):
			self.ui.label_17.setText("Wrong Password, Please Try Again.")
		elif not found:
			self.ui.label_17.setText("Username doesn't exist.")

	def comboBoxFixer(self):
		if self.ui.comboBox.isEditable() == True and self.ui.comboBox.currentText != "Choose Account Type":
			self.ui.comboBox.setCurrentText("Choose Account Type")


	def Register(self):
		username = self.ui.lineEdit_5.text()
		password = self.ui.lineEdit_6.text()
		confirm = self.ui.lineEdit_9.text()
		users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Admin','Patient','Nurse'])
		admins = users['Admin']
		patients = users['Patient']
		nurses = users['Nurse']
		requests = pd.read_excel("assets\\Database\\requests.xlsx",sheet_name=['Admin','Patient','Nurse'])
		chosen = self.ui.comboBox.currentText()
		found = False
		for user in [admins,patients,nurses,requests['Admin'],requests['Nurse'],requests['Patient']]:
			if found:
				break
			for i in user['username']:
				if username == i:
					found = True
					break
		if ' ' in username:
			self.ui.label_18.setText("Username can't contain spaces.")
		elif username == '' or password == '' or confirm == '' or chosen == "Choose Account Type":
			self.ui.label_18.setText("Please fill all requirements.")
		elif password != confirm:
			self.ui.label_18.setText("Passwords don't match.")
		elif found:
			self.ui.label_18.setText("Username is taken.")
		else:
			book = openpyxl.load_workbook('assets\\Database\\requests.xlsx')
			reqs = book[chosen]
			reqs.append([username,password])
			book.save(os.getcwd() + "\\assets\\Database\\requests.xlsx")
			self.ui.label_18.setText("Registered, waiting for admin confirmation.")
		
	def exit(self):
		self.close()

	def minimize(self):
		self.showNormal()
		self.showMinimized()


class MainWindow(QtWidgets.QMainWindow):
	def __init__(self):
		# modes: Admin , Patient , Nurse
		super().__init__()
		self.ui = uic.loadUi("GUI.ui", self)
		self.ui.full_menu.setVisible(False)
		self.ui.full_menu_2.setVisible(False)
		self.ui.full_menu_3.setVisible(False)
		self.ui.stackedWidget.setCurrentIndex(0)
		self.ui.stackedWidget_3.setCurrentIndex(0)
		self.ui.stackedWidget_4.setCurrentIndex(0)
		if session.username != 'H3SHAM' and session.username != 'h':
			self.ui.tabWidget.setTabVisible(2,False)
			self.ui.tabWidget_3.setTabVisible(2,False)
		if mode == 1:
			self.UpdateListWidgetsFromReqs([self.ui.listWidget,self.ui.listWidget_2,self.ui.listWidget_3],['Nurse','Patient','Admin'])
			self.UpdateListWidgetsFromUsers([self.ui.listWidget_6,self.ui.listWidget_7,self.ui.listWidget_8],['Nurse','Patient','Admin'])
			self.AdminUpdateTableFromMessages()		
			self.UpdateListWidgetsFromUsers([self.ui.listWidget_10],['Nurse'])
		elif mode == 2:
			self.UpdateListWidgetsFromUsers([self.ui.listWidget_4],['Nurse'])
			self.sw = Stopwatch()
			self.sw.start()
			self.ECGGraph = PlotWidget()
			self.x = []
			self.y = []
			layout1=QVBoxLayout()
			layout1.addWidget(self.ECGGraph)
			self.ui.widget_11.setLayout(layout1)
			self.data_line = self.ECGGraph.plot(pen='g',name='ECG')
			self.connected = True
			self.timepassed = 0
			ReadThread =threading.Thread(target= self.Reading)
			ReadThread.start()
			# try:
			# 	self.serialPort = serial.Serial(port='COM7', baudrate=9600)
			# except:
			# 	self.connected = False
			# else:
			# 	self.connected = True


		elif mode == 3:
			self.ui.lineEdit.setValidator(QtGui.QIntValidator(0,99,self))
			self.ui.lineEdit_4.setValidator(QtGui.QIntValidator(0,2147483647,self))
			self.UpdateListWidgetsFromUsers([self.ui.listWidget_5],['Patient'])
			self.UpdateListWidgetsFromUsers([self.ui.listWidget_9],['Patient'],'Yes')
			self.NurseUpdateTableFromMessages()

		self.ui.stackedWidget_2.setCurrentIndex(mode-1)

		self.ui.comboBox.currentIndexChanged.connect(self.UpdateCombo)
		self.ui.comboBox.currentTextChanged.connect(self.comboBoxFixer)
		self.ui.pushButton_9.clicked.connect(self.UpdateAdminSidebar)
		self.ui.pushButton_10.clicked.connect(self.UpdatePatientSidebar)
		self.ui.pushButton_11.clicked.connect(self.UpdateNurseSidebar)
		self.ui.Requests.clicked.connect(self.AdminReqsMode)
		self.ui.Requests2.clicked.connect(self.AdminReqsMode)
		self.ui.Statistics.clicked.connect(self.AdminStatsMode)
		self.ui.Statistics2.clicked.connect(self.AdminStatsMode)
		self.ui.Users.clicked.connect(self.AdminUsersMode)
		self.ui.Users2.clicked.connect(self.AdminUsersMode)
		self.ui.Technical.clicked.connect(self.AdminTechnicalMode)
		self.ui.Technical2.clicked.connect(self.AdminTechnicalMode)
		self.ui.Accept.clicked.connect(self.AcceptRequests)
		self.ui.PatReq.clicked.connect(self.NursePatReqsMode)
		self.ui.PatReq2.clicked.connect(self.NursePatReqsMode)
		self.ui.Tech.clicked.connect(self.NurseTechnicalMode)
		self.ui.Tech2.clicked.connect(self.NurseTechnicalMode)
		self.ui.PatProf.clicked.connect(self.NursePatProfMode)
		self.ui.PatProf2.clicked.connect(self.NursePatProfMode)
		self.ui.Statistics_2.clicked.connect(self.NurseStatsMode)
		self.ui.Statistics2_2.clicked.connect(self.NurseStatsMode)
		self.ui.Internet2.clicked.connect(self.PatientInternetMode)
		self.ui.Internet.clicked.connect(self.PatientInternetMode)
		self.ui.TV2.clicked.connect(self.PatientTVMode)
		self.ui.TV.clicked.connect(self.PatientTVMode)
		self.ui.NurseReq2.clicked.connect(self.PatientNurseReqMode)
		self.ui.NurseReq.clicked.connect(self.PatientNurseReqMode)
		self.ui.Vital2.clicked.connect(self.PatientVitalMode)
		self.ui.Vital.clicked.connect(self.PatientVitalMode)
		self.ui.exit.clicked.connect(self.LogOut)
		self.ui.exit2.clicked.connect(self.LogOut)
		self.ui.exit2_3.clicked.connect(self.LogOut)
		self.ui.exit_3.clicked.connect(self.LogOut)
		self.ui.exit2_2.clicked.connect(self.LogOut)
		self.ui.exit_2.clicked.connect(self.LogOut)
		self.ui.SaveChanges.clicked.connect(self.UpdatePatientInfo)
		self.ui.listWidget_5.itemSelectionChanged.connect(self.GetPatientInfo)
		self.ui.Delete.clicked.connect(self.DeleteUser)
		self.ui.PatRequest.clicked.connect(self.PatientReqNurse)
		self.ui.Done.clicked.connect(self.PatientReqDone)
		self.ui.Decline.clicked.connect(self.DeclineRequests)
		self.ui.Back.clicked.connect(self.WebBack)
		self.ui.Forward.clicked.connect(self.WebForward)
		self.ui.Reload.clicked.connect(self.WebReload)
		self.ui.SendReport.clicked.connect(self.NurseRepAdmin)
		self.ui.exit_4.clicked.connect(self.AdminRepNurse)
		self.ui.exit_5.clicked.connect(self.ClearAdminMessages)
		self.exit_6.clicked.connect(self.ClearNurseMessages)

	def Reading(self):
	# 	while self.connected == True:
	# 		data = self.serialPort.readline(1024)
	# 		if data:
	# 			data = str(data.decode('ascii'))
	# 			data = data.replace('\r\n','')
	# 			# print(data)
	# 			if data != '':
	# 				data = data.split('/')
	# 				ReadThread.returnValue = data
	# 				if len(data)>6:
	# 					if data[3] == '1':
	# 						self.PatientReqNurse()
	# 						print("Requested")
	# 					self.ui.label_5.setText(str(data[6]))
	# 					if data[5] != 'nan':
	# 						self.ui.label_7.setText(str(data[5]))
	# 						self.ui.label_8.setText(str(data[4]))
	# 			else:
	# 				return None
		while self.connected:
			if self.sw.secondsPassed() > self.timepassed:
				self.ui.label_7.setText(str(random.randint(24,27)))
				self.ui.label_8.setText(str(random.randint(20,30)))
				ecg = random.randint(65,75)
				self.ui.label_5.setText(str(ecg))
				self.x.append(self.timepassed)
				self.y.append(ecg)
				self.data_line.setData(self.x,self.y)
				self.timepassed += 1
				

	def comboBoxFixer(self):
		if self.ui.comboBox.isEditable() == True and self.ui.comboBox.currentText != "Gender":
			self.ui.comboBox.setCurrentText("Gender")

	def UpdateCombo(self):
		self.ui.comboBox.setEditable(False)

	def WebBack(self):
		self.ui.webEngineView.back()

	def WebForward(self):
		self.ui.webEngineView.forward()

	def WebReload(self):
		self.ui.webEngineView.reload()

	def GetPatientInfo(self):
		users = openpyxl.load_workbook("assets\\Database\\users.xlsx")
		patients = users["Patient"]
		acceptIndices = []
		tab = self.ui.tabWidget_2.currentWidget().findChildren(QtWidgets.QListWidget)[0]
		for i in tab.selectionModel().selectedIndexes():
			acceptIndices.append(i.row())
		acceptIndices.sort(reverse=True)
		for i in acceptIndices:
			self.ui.lineEdit_2.setText(patients.cell(row=i+2,column=5).value)
			self.ui.lineEdit_4.setText(patients.cell(row=i+2,column=6).value)
			self.ui.lineEdit.setText(patients.cell(row=i+2,column=3).value)
			if patients.cell(row=i+2,column=4).value == 'Male':
				self.ui.comboBox.setCurrentIndex(0)
			elif patients.cell(row=i+2,column=4).value == 'Female':
				self.ui.comboBox.setCurrentIndex(1)

	def UpdatePatientInfo(self):
		Email = self.ui.lineEdit_2.text()
		Number = self.ui.lineEdit_4.text()
		Age = self.ui.lineEdit.text()
		Gender = self.ui.comboBox.currentText()
		if ('@' not in Email or '.com' not in Email) or Email == '':
			self.ui.label_4.setText("Please enter a valid email.")
		elif Number == '' or int(Number) < 1000000000 or int(Number) > 1500000000:
			self.ui.label_4.setText("Please enter a valid mobile number.")
		elif Age == '':
			self.ui.label_4.setText("Please enter age.")
		elif Gender == 'Gender':
			self.ui.label_4.setText("Please choose gender.")
		else:
			acceptIndices = []
			tab = self.ui.tabWidget_2.currentWidget().findChildren(QtWidgets.QListWidget)[0]
			for i in tab.selectionModel().selectedIndexes():
				acceptIndices.append(i.row())
			users = openpyxl.load_workbook("assets\\Database\\users.xlsx")
			patients = users["Patient"]
			for i in acceptIndices:
				patients.cell(row=i+2,column=3).value = Age
				patients.cell(row=i+2,column=4).value = Gender
				patients.cell(row=i+2,column=5).value = Email
				patients.cell(row=i+2,column=6).value = Number
			users.save(os.getcwd() + "\\assets\\Database\\users.xlsx")

	def AcceptRequests(self):
		acceptIndices = []
		tab = self.ui.tabWidget.currentWidget().findChildren(QtWidgets.QListWidget)[0]
		for i in tab.selectionModel().selectedIndexes():
			acceptIndices.append(i.row())
		acceptIndices.sort(reverse=True)
		profession = self.ui.tabWidget.tabText(self.ui.tabWidget.currentIndex())
		reqs = openpyxl.load_workbook('assets\\Database\\requests.xlsx')
		users = openpyxl.load_workbook('assets\\Database\\users.xlsx')
		needed = reqs[profession]
		dest = users[profession]
		for i in acceptIndices:
			username = needed.cell(row=i+2,column=1).value
			password = needed.cell(row=i+2,column=2).value
			needed.delete_rows(i+2)
			dest.append([username,password])
			reqs.save(os.getcwd() + "\\assets\\Database\\requests.xlsx")
			users.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
		self.UpdateListWidgetsFromReqs([self.ui.listWidget,self.ui.listWidget_2,self.ui.listWidget_3],['Nurse','Patient','Admin'])
		self.UpdateListWidgetsFromUsers([self.ui.listWidget_6,self.ui.listWidget_7,self.ui.listWidget_8],['Nurse','Patient','Admin'])
		self.UpdateListWidgetsFromUsers([self.ui.listWidget_10],['Nurse'])

	def DeclineRequests(self):
		acceptIndices = []
		tab = self.ui.tabWidget.currentWidget().findChildren(QtWidgets.QListWidget)[0]
		for i in tab.selectionModel().selectedIndexes():
			acceptIndices.append(i.row())
		acceptIndices.sort(reverse=True)
		profession = self.ui.tabWidget.tabText(self.ui.tabWidget.currentIndex())
		reqs = openpyxl.load_workbook('assets\\Database\\requests.xlsx')
		needed = reqs[profession]
		for i in acceptIndices:
			username = needed.cell(row=i+2,column=1).value
			password = needed.cell(row=i+2,column=2).value
			needed.delete_rows(i+2)
			reqs.save(os.getcwd() + "\\assets\\Database\\requests.xlsx")
			self.UpdateListWidgetsFromReqs([self.ui.listWidget,self.ui.listWidget_2,self.ui.listWidget_3],['Nurse','Patient','Admin'])

	def NurseRepAdmin(self):
		title = self.ui.lineEdit_3.text()
		message = self.ui.plainTextEdit.toPlainText()
		now = datetime.now()
		dt = now.strftime("%d/%m/%Y %H:%M:%S")
		x = dt.split(' ')
		date = x[0]
		time = x[1]
		username = session.username
		receiver = "Admin"
		if title == '':
			self.ui.label_11.setText("Please insert title.")
		elif message == '':
			self.ui.label_11.setText("Please insert message")
		else:
			messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
			msg = messages['Message']
			msg.append([username,receiver,date,time,title,message])
			messages.save(os.getcwd() + "\\assets\\Database\\messages.xlsx")
			self.ui.label_11.setText("Message successfully sent.")

	def AdminRepNurse(self):
		title = self.ui.lineEdit_5.text()
		message = self.ui.plainTextEdit_3.toPlainText()
		now = datetime.now()
		dt = now.strftime("%d/%m/%Y %H:%M:%S")
		x = dt.split(' ')
		date = x[0]
		time = x[1]
		username = session.username
		tab = self.ui.listWidget_10
		receiver = tab.currentItem().text()
		if title == '':
			self.ui.label_13.setText("Please insert title.")
		elif message == '':
			self.ui.label_13.setText("Please insert message.")
		elif receiver == '':
			self.ui.label_13.setText("Please choose a nurse.")
		else:
			messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
			msg = messages['Message']
			msg.append([username,receiver,date,time,title,message])
			messages.save(os.getcwd() + "\\assets\\Database\\messages.xlsx")
			self.ui.label_13.setText("Message successfully sent.")


	def AdminUpdateTableFromMessages(self):
		table = self.ui.tableWidget_3
		messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
		msg = messages["Message"]
		msgs = pd.read_excel("assets\\Database\\messages.xlsx",sheet_name=['Message'])
		c = []
		co = 0
		for i in msgs["Message"]['receiver']:
			if i == "Admin":
				c.append(co)
			co+=1
		table.setRowCount(len(c))
		for i,ii in zip(c,range(len(c))):
			table.setItem(ii,0,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=1).value))
			table.setItem(ii,1,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=3).value))
			table.setItem(ii,2,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=4).value))
			table.setItem(ii,3,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=5).value))
			table.setItem(ii,4,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=6).value))
			table.setItem(ii,5,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=7).value))
		
	def NurseUpdateTableFromMessages(self):
		table = self.ui.tableWidget
		messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
		msg = messages["Message"]
		msgs = pd.read_excel("assets\\Database\\messages.xlsx",sheet_name=['Message'])
		c = []
		co = 0
		for i in msgs["Message"]['receiver']:
			if i == session.username:
				c.append(co)
			co+=1
		table.setRowCount(len(c))
		for i,ii in zip(c,range(len(c))):
			table.setItem(ii,0,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=1).value))
			table.setItem(ii,1,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=3).value))
			table.setItem(ii,2,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=4).value))
			table.setItem(ii,3,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=5).value))
			table.setItem(ii,4,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=6).value))
			table.setItem(ii,5,QtWidgets.QTableWidgetItem(msg.cell(row=i+2,column=7).value))

	def ClearAdminMessages(self):
		table = self.ui.tableWidget_3
		messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
		msg = messages["Message"]
		msgs = pd.read_excel("assets\\Database\\messages.xlsx",sheet_name=['Message'])
		co = 0
		c = []
		for i in msgs["Message"]['receiver']:
			if i == "Admin":
				c.append(co)
			co+=1
		for i in c:
			msg.delete_rows(i+2,1)
			messages.save(os.getcwd() + "\\assets\\Database\\messages.xlsx")
		table.clearContents()
		table.setRowCount(0)

	def ClearNurseMessages(self):
		table = self.ui.tableWidget
		messages = openpyxl.load_workbook("assets\\Database\\messages.xlsx")
		msg = messages["Message"]
		msgs = pd.read_excel("assets\\Database\\messages.xlsx",sheet_name=['Message'])
		co = 0
		c = []
		for i in msgs["Message"]['receiver']:
			if i == session.username:
				c.append(co)
			co+=1
		for i in c:
			msg.delete_rows(i+2,1)
			messages.save(os.getcwd() + "\\assets\\Database\\messages.xlsx")
		table.clearContents()
		table.setRowCount(0)

	def UpdateListWidgetsFromReqs(self,listWidgets,accounts):
		requests = pd.read_excel("assets\\Database\\requests.xlsx",sheet_name=['Admin','Patient','Nurse'])
		for type,i in zip(accounts,listWidgets):
			i.clear()
			for user in [requests[type]]: 
				for j in user['username']:
					i.addItem(j)
		
	def UpdateListWidgetsFromUsers(self,listWidgets,accounts,reqflag=None):
		requests = pd.read_excel("assets\\Database\\users.xlsx",sheet_name=['Admin','Patient','Nurse'])
		reqs = openpyxl.load_workbook("assets\\Database\\users.xlsx")
		c = 1
		for type,i in zip(accounts,listWidgets):
			i.clear()
			for user in [requests[type]]: 
				for j in user['username']:
					c+=1
					if j != "H3SHAM" and j != 'h':
						if reqflag == None:
							i.addItem(j)
						elif reqflag == 'Yes':
							if reqs['Patient'].cell(row=c,column=7).value == 'Yes':
								i.addItem(j)
							
	def PatientReqDone(self):
		acceptIndices = []
		tab = self.ui.listWidget_9
		acceptIndices = tab.currentItem().text()
		users = openpyxl.load_workbook('assets\\Database\\users.xlsx')
		needed = users['Patient']
		users1 = pd.read_excel("assets\\Database\\users.xlsx",sheet_name=['Admin','Patient','Nurse'])
		row = 0
		c = 1
		for i in users1['Patient']['username']:
			c += 1
			if i == acceptIndices:
				row = c
				break
		for i in acceptIndices:
			needed.cell(row=row,column=7).value = ''
			users.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
		self.UpdateListWidgetsFromUsers([self.ui.listWidget_9],['Patient'],'Yes')

	def DeleteUser(self):
		acceptIndices = []
		tab = self.ui.tabWidget_3.currentWidget().findChildren(QtWidgets.QListWidget)[0]
		for i in tab.selectionModel().selectedIndexes():
			acceptIndices.append(i.row())
		acceptIndices.sort(reverse=True)
		profession = self.ui.tabWidget.tabText(self.ui.tabWidget_3.currentIndex())
		users = openpyxl.load_workbook('assets\\Database\\users.xlsx')
		needed = users[profession]
		c=2
		if profession == "Admin":
			c=4
		for i in acceptIndices:
			needed.delete_rows(i+c)
			users.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
		self.UpdateListWidgetsFromUsers([self.ui.listWidget_6,self.ui.listWidget_7,self.ui.listWidget_8],['Nurse','Patient','Admin'])
		self.UpdateListWidgetsFromUsers([self.ui.listWidget_10],['Nurse'])

	def PatientReqNurse(self):
		users = openpyxl.load_workbook('assets\\Database\\users.xlsx')
		Patients = users['Patient']
		users1 = pd.read_excel("assets\\Database\\users.xlsx",sheet_name=['Admin','Patient','Nurse'])
		row = 0
		c = 1
		for i in users1['Patient']['username']:
			c += 1
			if i == session.username:
				row = c
				break
		Patients.cell(row=c,column=7).value = 'Yes'
		users.save(os.getcwd() + "\\assets\\Database\\users.xlsx")


	def UpdateAdminSidebar(self):
		if self.ui.pushButton_9.isChecked() == True:
			self.ui.Requests2.setHidden(True)
			self.ui.Statistics2.setHidden(True)
			self.ui.exit2.setHidden(True)
			self.ui.Users2.setHidden(True)
			self.ui.Technical2.setHidden(True)
		else:
			self.ui.Requests2.setHidden(False)
			self.ui.Statistics2.setHidden(False)
			self.ui.exit2.setHidden(False)
			self.ui.Users2.setHidden(False)
			self.ui.Technical2.setHidden(False)

	def UpdatePatientSidebar(self):
		if self.ui.pushButton_10.isChecked() == True:
			self.ui.exit2_2.setHidden(True)
			self.ui.Internet2.setHidden(True)
			self.ui.TV2.setHidden(True)
			self.ui.NurseReq2.setHidden(True)
			self.ui.Vital2.setHidden(True)
		else:
			self.ui.exit2_2.setHidden(False)
			self.ui.Internet2.setHidden(False)
			self.ui.TV2.setHidden(False)
			self.ui.NurseReq2.setHidden(False)
			self.ui.Vital2.setHidden(False)

	def UpdateNurseSidebar(self):
		if self.ui.pushButton_11.isChecked() == True:
			self.ui.exit2_3.setHidden(True)
			self.ui.PatReq2.setHidden(True)
			self.ui.Tech2.setHidden(True)
			self.ui.PatProf2.setHidden(True)
			self.ui.Statistics2_2.setHidden(True)
		else:
			self.ui.exit2_3.setHidden(False)
			self.ui.PatReq2.setHidden(False)
			self.ui.Tech2.setHidden(False)
			self.ui.PatProf2.setHidden(False)
			self.ui.Statistics2_2.setHidden(False)


	def AdminReqsMode(self):
		self.ui.stackedWidget.setCurrentIndex(0)

	def AdminStatsMode(self):
		self.ui.stackedWidget.setCurrentIndex(2)

	def AdminUsersMode(self):
		self.ui.stackedWidget.setCurrentIndex(1)

	def AdminTechnicalMode(self):
		self.ui.stackedWidget.setCurrentIndex(3)

	def NursePatReqsMode(self):
		self.ui.stackedWidget_4.setCurrentIndex(0)

	def NurseTechnicalMode(self):
		self.ui.stackedWidget_4.setCurrentIndex(1)

	def NursePatProfMode(self):
		self.ui.stackedWidget_4.setCurrentIndex(2)

	def NurseStatsMode(self):
		self.ui.stackedWidget_4.setCurrentIndex(3)

	def PatientInternetMode(self):
		self.ui.stackedWidget_3.setCurrentIndex(0)
	
	def PatientTVMode(self):
		self.ui.stackedWidget_3.setCurrentIndex(1)

	def PatientNurseReqMode(self):
		self.ui.stackedWidget_3.setCurrentIndex(2)
	
	def PatientVitalMode(self):
		self.ui.stackedWidget_3.setCurrentIndex(3)

	def LogOut(self):
		win = LoginWindow()
		win.setWindowTitle("Sign Up")
		win.setWindowIcon(QtGui.QIcon("assets\\865969.png"))
		win.show()
		# self.serialPort.close()
		self.close()
		
	def closeEvent(self, a0: QtGui.QCloseEvent):
		# self.serialPort.close()
		self.connected = False
		return super().closeEvent(a0)